from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook


# Builds a workbook with a bunch of values to do math on
values_workbook = Workbook()

dest_filename = 'some_good_data.xlsx'

worksheet1 = values_workbook.active
worksheet1.title = "range names"

for row in range(1, 40):
    worksheet1.append(range(600))

i = 0
for index in range(1,255):
    if index % 2 == 0:
        worksheet1.cell(column = index, row = 1, value="Header "+`index`)
    else:
        worksheet1.cell(column = index, row = 1, value=`index`)


worksheet2 = values_workbook.create_sheet(title="Pi")

worksheet2['F5'] = 3.14

worksheet3 = values_workbook.create_sheet(title="Data")
for row in range(10, 20):
    for col in range(27, 54):
        _ = worksheet3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))

values_workbook.save(filename = dest_filename)
# End building excel sheet with lots of values

# Read in excel sheet built earlier
values_workbook_read = load_workbook(filename = 'some_good_data.xlsx')
# get worksheet named range names
worksheet1 = values_workbook_read['range names']

# add all of the 5th row numbers together where column 1 contains the string Header
output_number = 0
for index, row in enumerate(worksheet1.iter_rows()):
    if 'Header' in worksheet1.cell(row=1, column=index + 1).value:
        #print(worksheet1.cell(row=1, column= index + 1).value) 
        output_number += worksheet1.cell(row=5, column=index+1).value

print(output_number)